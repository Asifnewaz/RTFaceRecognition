import pickle

import face_recognition
from keras.src.models import Sequential
from keras.src.layers import Conv2D, MaxPooling2D, Flatten, Dense
from keras.src.saving import load_model
from keras.src.utils import to_categorical
import cv2
import numpy as np
from sklearn.model_selection import train_test_split
from PyQt5.QtWidgets import QApplication, QLabel, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QMessageBox, QStackedWidget
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer, QThread, pyqtSignal, Qt
import os
import time
import openpyxl
from datetime import datetime

class CountdownThread(QThread):
    countdown_signal = pyqtSignal(int)

    def run(self):
        for i in range(3, 0, -1):
            self.countdown_signal.emit(i)
            self.sleep(1)

class FaceRecognitionApp(QWidget):
    def __init__(self):
        super().__init__()

        # Initialize main stacked widget
        self.time_label = None
        self.id_label = None
        self.ready_button = None
        self.ready_button_clicked = False

        self.stack = QStackedWidget()
        self.setWindowTitle("Face Recognition App")
        self.setGeometry(100, 100, 800, 600)

        # Create pages
        self.start_page = self.create_start_page()
        self.video_feed_page = self.create_video_feed_page()
        self.countdown_label = QLabel("", self)

        # Add pages to stack
        self.stack.addWidget(self.start_page)
        self.stack.addWidget(self.video_feed_page)

        # Set layout
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.stack)
        self.setLayout(main_layout)

        # Load the encoding file
        print("Loading Encode File ...")
        file = open('EncodeFile.p', 'rb')
        encode_list_known_with_ids = pickle.load(file)
        file.close()
        self.encodeListKnown, self.studentIds = encode_list_known_with_ids


        self.detected_faces = set()
        self.pause_processing = False  # Initialize pause_processing
        self.processing = False  # Initialize processing
        self.cap = None
        self.timer = None

    def create_start_page(self):
        start_page = QWidget()
        layout = QVBoxLayout()

        # Class ID Input
        self.class_id_input = QLineEdit()
        self.class_id_input.setPlaceholderText("Enter Class ID")
        self.class_id_input.setFixedWidth(350)  # Set width to 350 px
        self.class_id_input.setStyleSheet(
            "font-size: 16px; padding: 5px; margin-bottom: 10px; border: 2px dotted gray; border-radius: 5px;"
        )  # Added dotted border and set width
        layout.addWidget(self.class_id_input, alignment=Qt.AlignCenter)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()

        # Start Button
        start_button = QPushButton("Start attendance system")
        start_button.setStyleSheet(
            "font-size: 16px; padding: 10px; color: green; border: 2px dotted green; border-radius: 5px; margin-right: 10px;"
        )  # Added dotted border and text color
        start_button.clicked.connect(self.start_countdown)
        button_layout.addWidget(start_button)

        # Add another button horizontally
        stop_button = QPushButton("Add new person")
        stop_button.setStyleSheet(
            "font-size: 16px; padding: 10px; color: red; border: 2px dotted red; border-radius: 5px; margin-left: 10px;"
        )  # Added dotted border and text color
        stop_button.clicked.connect(self.close)  # Example functionality to close the app
        button_layout.addWidget(stop_button)

        # Center-align the buttons
        button_layout.setAlignment(Qt.AlignCenter)

        # Add the horizontal button layout to the main vertical layout
        layout.addLayout(button_layout)

        start_page.setLayout(layout)
        return start_page

    def create_video_feed_page(self):
        video_feed_page = QWidget()
        layout = QHBoxLayout()

        # Camera Feed Layout (Left Side)
        video_layout = QVBoxLayout()
        self.video_feed = QLabel()
        self.video_feed.setFixedSize(500, 375)  # Set video feed width to 500px
        video_layout.addWidget(self.video_feed)
        layout.addLayout(video_layout)

        # Controls and Info Layout (Right Side)
        controls_layout = QVBoxLayout()

        # Spacer for top alignment
        controls_layout.addStretch()

        # Ready Button Right Centered
        self.ready_button = QPushButton("Ready")
        self.ready_button.setStyleSheet(
            "font-size: 16px; padding: 10px; color: white; background-color: green; border: none; border-radius: 5px;"
        )
        self.ready_button.clicked.connect(self.enable_processing)
        controls_layout.addWidget(self.ready_button, alignment=Qt.AlignVCenter | Qt.AlignRight)

        # Spacer for bottom alignment
        controls_layout.addStretch()

        # ID Label
        self.id_label = QLabel("ID: ")
        self.id_label.setStyleSheet("font-size: 16px;")
        self.id_label.setVisible(False)
        controls_layout.addWidget(self.id_label)

        # Time Label
        self.time_label = QLabel("Time: ")
        self.time_label.setStyleSheet("font-size: 16px;")
        self.time_label.setVisible(False)
        controls_layout.addWidget(self.time_label)

        # Spacer for alignment
        controls_layout.addStretch()

        layout.addLayout(controls_layout)
        video_feed_page.setLayout(layout)

        return video_feed_page

    def enable_processing(self):
        # Enable image processing when the "Ready" button is clicked
        self.ready_button.setVisible(False)  # Hide the button
        self.pause_processing = False  # Allow processing to start
        self.ready_button_clicked = True

    def start_countdown(self):
        self.class_id = self.class_id_input.text().strip()
        if not self.class_id:
            QMessageBox.warning(self, "Error", "Please enter a valid Class ID.")
            return

        self.countdown_label.setText("Starting in...")
        self.countdown_label.setAlignment(Qt.AlignCenter)
        self.countdown_label.setStyleSheet("font-size: 24px; color: red;")
        self.stack.addWidget(self.countdown_label)
        self.stack.setCurrentWidget(self.countdown_label)

        self.countdown_thread = CountdownThread()
        self.countdown_thread.countdown_signal.connect(self.update_countdown)
        self.countdown_thread.finished.connect(self.start_video_feed)
        self.countdown_thread.start()

    def update_countdown(self, value):
        messages = {3: "Preparing data...", 2: "Loading data...", 1: "Fetching data..."}
        if value in messages:
            full_message = messages[value]
            self.text_animation_index = 0  # Reset animation index
            self.text_animation_timer = QTimer()
            self.text_animation_timer.timeout.connect(lambda: self.animate_text(full_message))
            self.text_animation_timer.start(100)  # Adjust timing for each letter

    def animate_text(self, full_message):
        if self.text_animation_index <= len(full_message):
            self.countdown_label.setText(full_message[:self.text_animation_index])
            self.text_animation_index += 1
        else:
            self.text_animation_timer.stop()

    def start_video_feed(self):
        self.stack.setCurrentWidget(self.video_feed_page)

        self.cap = cv2.VideoCapture(0)
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_feed)
        self.timer.start(10)

    def update_feed(self):

        ret, frame = self.cap.read()

        if ret:
            # Resize frame to center the face
            height, width, _ = frame.shape
            crop_size = min(height, width)
            start_x = (width - crop_size) // 2
            start_y = (height - crop_size) // 2
            frame = frame[start_y:start_y + crop_size, start_x:start_x + crop_size]
            frame = cv2.resize(frame, (640, 480))

            # Update video feed
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            bytes_per_line = ch * w
            qimg = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            self.video_feed.setPixmap(QPixmap.fromImage(qimg))

            if self.pause_processing:
                return

            if not self.ready_button_clicked:
                return

            # If not processing, trigger detection
            if not self.processing:
                self.processing = True
                QTimer.singleShot(500, lambda: self.handle_detection(frame))

    def handle_detection(self, frame):
        self.pause_processing = True
        result = self.process_frame(frame)

        if result == "Unknown":
            # Clear labels if no face is detected
            self.id_label.setText("ID: ")
            self.time_label.setText("Time: ")
        else:
            self.id_label.setText(f"ID: {result}")
            self.time_label.setText(f"Time: {time.strftime('%H:%M:%S')}")
            self.id_label.setVisible(True)
            self.time_label.setVisible(True)

            if result in self.detected_faces:
                msg = QMessageBox()
                msg.setWindowTitle("Attendance Saved")
                msg.setText("User already marked present.")
                msg.setStandardButtons(QMessageBox.Ok)
                response = msg.exec_()  # Capture the response
                if response == QMessageBox.Ok:
                    self.reset_data()
            else:
                self.detected_faces.add(result)
                self.save_to_spreadsheet(result)

                msg = QMessageBox()
                msg.setWindowTitle("Recognition Success")
                msg.setText("User recognized successfully!")
                msg.setStandardButtons(QMessageBox.Ok)
                response = msg.exec_()  # Capture the response
                if response == QMessageBox.Ok:
                    self.reset_data()

        self.processing = False
        self.pause_processing = False

    def reset_data(self):
        print("OK button was clicked.")
        self.ready_button.setVisible(True)
        self.ready_button_clicked = False
        self.id_label.setVisible(False)
        self.time_label.setVisible(False)

    def save_to_spreadsheet(self, user_id):
        # Define the spreadsheet path
        file_path = "attendance.xlsx"

        # Load or create the workbook
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()

        sheet = workbook.active
        # Check if the headers exist
        if sheet.max_row == 1:
            sheet.append(["ID", "Name", "Time"])

        # Append the data (ID, Name, and current time)
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([user_id, self.get_name_from_id(user_id), current_time])

        # Save the workbook
        workbook.save(file_path)

    def get_name_from_id(self, user_id):
        # You can define a method to get the name based on the ID,
        # you could have a dictionary or another method to map IDs to names
        # Example:
        name_mapping = {
            '1': 'John Doe',
            '2': 'Jane Smith',
            # Add other IDs and names
        }
        return name_mapping.get(user_id, "Unknown")

    def process_frame(self, image):
        # Convert BGR (OpenCV) to RGB (face_recognition library)
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # # Find all face locations in the frame
        face_locations = face_recognition.face_locations(rgb_image)
        encode_cur_frame = face_recognition.face_encodings(rgb_image, face_locations)

        result = "Unknown"

        if face_locations:
            for encodeFace, faceLoc in zip(encode_cur_frame, face_locations):
                matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)
                # print("matches", matches)
                # print("faceDis", faceDis)

                match_index = np.argmin(faceDis)
                # print("Match Index", matchIndex)

                if matches[match_index]:
                    # print("Known Face Detected")
                    self.id = self.studentIds[match_index]
                    top, right, bottom, left = faceLoc
                    cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)

                    # Display text (e.g., 'Face Detected') at the top-left corner of the face
                    result =   self.id
                print(self.id)

        return result

    def closeEvent(self, event):
        if self.cap:
            self.cap.release()
        event.accept()

if __name__ == "__main__":
    app = QApplication([])
    window = FaceRecognitionApp()
    window.show()
    app.exec_()