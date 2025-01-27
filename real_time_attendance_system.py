import pickle
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage
import face_recognition
from keras.src.models import Sequential
from keras.src.layers import Conv2D, MaxPooling2D, Flatten, Dense
from keras.src.saving import load_model
from keras.src.utils import to_categorical
import cv2
import numpy as np
from sklearn.model_selection import train_test_split
from PyQt5.QtWidgets import QApplication, QLabel, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QMessageBox, QStackedWidget
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import QTimer, QThread, pyqtSignal, Qt
import os
import time
import openpyxl
from datetime import datetime
from AddPersonPage import AddPersonPage

cred = credentials.Certificate("secrets/serviceAccountKey2.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': "https://faceattendancerealtime-9e847-default-rtdb.firebaseio.com/"
})


class CountdownThread(QThread):
    countdown_signal = pyqtSignal(int)

    def run(self):
        for i in range(3, 0, -1):
            self.countdown_signal.emit(i)
            self.sleep(1)

class FaceRecognitionApp(QWidget):
    def __init__(self):
        super().__init__()

        self.id = None
        self.encodeListKnown = None
        self.studentIds = None
        self.attendance_label = None
        self.date_label = None
        self.name_label = None
        self.video_feed = None
        self.windowTitle = "Face Recognition"
        self.windowWidth = 1200
        self.windowHeight = 800
        self.windowBackgroundColor = "background-color: #d8dee9;"

        self.videoFeedWidth = 800
        self.videoFeedHeight = 533
        self.right_panelWidth = 340
        self.right_panelHeight = 760

        self.padding = 20
        self.rightPadding = 20
        self.topPadding = 20

        # Initialize main stacked widget
        self.time_label = None
        self.id_label = None
        self.ready_button = None
        self.back_button = None
        self.ready_button_clicked = False

        self.stack = QStackedWidget()
        self.setWindowTitle(self.windowTitle)
        self.setGeometry(0,0, self.windowWidth, self.windowHeight)
        self.setStyleSheet(self.windowBackgroundColor)  # Set window background color

        # Create pages
        self.start_page = self.create_start_page()
        self.video_feed_page = self.create_video_feed_page()
        self.countdown_label = QLabel("", self)

        # Add pages to stack
        self.stack.addWidget(self.start_page)
        self.stack.addWidget(self.video_feed_page)

        # Set layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)  # Remove default layout margins
        main_layout.setSpacing(0)  # Remove spacing between widgets
        main_layout.addWidget(self.stack)
        self.setLayout(main_layout)

        self.detected_faces = set()
        self.pause_processing = False  # Initialize pause_processing
        self.processing = False  # Initialize processing
        self.cap = None
        self.timer = None

        # Load CNN model
        # self.model_path = "pretrained_face_model.h5"
        # self.model = self.load_model()
        # self.class_ids = ""  # To store class ID
        # self.detected_facess = set()


    def load_model(self):
        try:
            model = load_model(self.model_path, compile=False)
            model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])
            return model
        except Exception as e:
            print(f"Failed to load model: {e}")
            exit()

    def create_start_page(self):
        start_page = QWidget()
        layout = QVBoxLayout()

        # Class ID Input
        self.class_id_input = QLineEdit()
        self.class_id_input.setPlaceholderText("Enter Class ID")
        self.class_id_input.setFixedWidth(350)  # Set width to 350 px
        self.class_id_input.setStyleSheet(
            "font-size: 16px; padding: 5px; margin-bottom: 10px; border: 2px dotted gray; border-radius: 5px;"
        )  # Added dotted border and set width
        layout.addWidget(self.class_id_input, alignment=Qt.AlignCenter)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()

        # Start Button
        start_button = QPushButton("Start attendance system")
        start_button.setStyleSheet(
            "font-size: 16px; padding: 10px; color: green; border: 2px dotted green; border-radius: 5px; margin-right: 10px;"
        )  # Added dotted border and text color
        start_button.clicked.connect(self.start_countdown)
        button_layout.addWidget(start_button)

        # Add another button horizontally
        add_new_person = QPushButton("Add new person")
        add_new_person.setStyleSheet(
            "font-size: 16px; padding: 10px; color: red; border: 2px dotted red; border-radius: 5px; margin-left: 10px;"
        )  # Added dotted border and text color
        add_new_person.clicked.connect(self.open_add_person_page)  # Example functionality to close the app
        button_layout.addWidget(add_new_person)

        # Center-align the buttons
        button_layout.setAlignment(Qt.AlignCenter)

        # Add the horizontal button layout to the main vertical layout
        layout.addLayout(button_layout)

        start_page.setLayout(layout)
        return start_page

    def create_video_feed_page(self):
        video_feed_page = QWidget()
        video_feed_page.setStyleSheet("background-color: #d8dee9;")  # Set window background color

        # Camera Feed (Left Panel)
        self.video_feed = QLabel(video_feed_page)
        self.video_feed.setGeometry(self.padding, 133, self.videoFeedWidth, self.videoFeedHeight)  # Left: 20, Top: 133, Width: 800, Height: 533
        self.video_feed.setStyleSheet("background-color: #ffffff; border: 1px solid #b0bec5;")

        # Right Panel
        right_panel = QWidget(video_feed_page)
        right_panel.setGeometry(self.windowWidth - (self.right_panelWidth + self.padding), self.padding, self.right_panelWidth, self.right_panelHeight)  # Right: 860, Top: 20, Width: 340, Height: 760
        right_panel.setStyleSheet("background-color: #e5e9f0; border: 0px solid #b0bec5; border-radius: 8px;")

        # ID Label
        self.id_label = QLabel("ID: ", right_panel)
        self.id_label.setGeometry(self.padding, 113, 300, 30)
        self.id_label.setStyleSheet("font-size: 18px;")

        # Name Label
        self.name_label = QLabel("Name: ", right_panel)
        self.name_label.setGeometry(self.padding, 163, 300, 30)
        self.name_label.setStyleSheet("font-size: 18px;")

        # Time Label
        self.time_label = QLabel("Time: ", right_panel)
        self.time_label.setGeometry(self.padding, 203, 300, 30)
        self.time_label.setStyleSheet("font-size: 18px;")

        # Date Label
        self.date_label = QLabel("Date: ", right_panel)
        self.date_label.setGeometry(self.padding, 243, 300, 30)
        self.date_label.setStyleSheet("font-size: 18px;")

        # Attendance Status Label
        self.attendance_label = QLabel("Attendance Marked", right_panel)
        self.attendance_label.setGeometry(self.padding, 620, 300, 30)
        self.attendance_label.setStyleSheet("font-size: 18px; color: green;")

        # Ready Button
        self.ready_button = QPushButton("Ready", right_panel)
        self.ready_button.setGeometry((self.right_panelWidth - 100) // 2, (self.right_panelHeight - 40) // 2, 100,
                                      40)  # Center the button
        self.ready_button.setStyleSheet(
            "font-size: 18px; padding: 10px; color: white; background-color: green; border: none; border-radius: 8px;"
        )
        self.ready_button.clicked.connect(self.enable_processing)

        self.back_button = QPushButton("Back", right_panel)
        self.back_button.setGeometry((self.right_panelWidth - 100) // 2, (self.right_panelHeight + 220) // 2, 100,
                                      40)  # Center the button
        self.back_button.setStyleSheet(
            "font-size: 18px; padding: 10px; color: white; background-color: red; border: none; border-radius: 8px;"
        )
        self.back_button.clicked.connect(self.back_to_initial_state)

        self.reset_data()
        return video_feed_page

    def back_to_initial_state(self):
        # Stop the camera and release resources
        self.ready_button_clicked = False
        self.encodeListKnown = None
        self.studentIds = None
        if self.cap and self.cap.isOpened():
            self.cap.release()

        # Stop the timer if running
        if self.timer:
            self.timer.stop()

        # Reset to start page
        self.stack.setCurrentWidget(self.start_page)

    def enable_processing(self):
        # Enable image processing when the "Ready" button is clicked
        self.ready_button.setVisible(False)  # Hide the button
        self.back_button.setVisible(False)  # Hide the button
        self.pause_processing = False  # Allow processing to start
        self.ready_button_clicked = True

    def start_countdown(self):
        self.class_id = self.class_id_input.text().strip()
        if not self.class_id:
            QMessageBox.warning(self, "Error", "Please enter a valid Class ID.")
            return

        self.countdown_label.setText("Starting in...")
        self.countdown_label.setAlignment(Qt.AlignCenter)
        self.countdown_label.setStyleSheet("font-size: 24px; color: red;")
        self.stack.addWidget(self.countdown_label)
        self.stack.setCurrentWidget(self.countdown_label)

        self.countdown_thread = CountdownThread()
        self.countdown_thread.countdown_signal.connect(self.update_countdown)
        self.countdown_thread.finished.connect(self.validate_and_start_video_feed)
        self.countdown_thread.start()

    def validate_and_start_video_feed(self):
        if self.encodeListKnown is not None:
            self.start_video_feed()
        else:
            self.stack.setCurrentWidget(self.start_page)

    def update_countdown(self, value):
        messages = {3: "Preparing data ...", 2: "Loading data ...", 1: "Fetching data ..."}
        if value in messages:
            full_message = messages[value]
            self.text_animation_index = 0  # Reset animation index
            self.text_animation_timer = QTimer()
            self.text_animation_timer.timeout.connect(lambda: self.animate_text(full_message))
            if value == 1:
                print("Loading Encode File ...")
                class_id = self.class_id_input.text().strip()
                fileName = class_id + "_model.p"
                if os.path.exists(fileName):
                    with open(fileName, 'rb') as file:
                        try:
                            encode_list_known_with_ids = pickle.load(file)
                            self.encodeListKnown, self.studentIds = encode_list_known_with_ids
                        except (pickle.UnpicklingError, EOFError) as e:
                            QMessageBox.critical(self, "Error", f"Error loading file: {str(e)}")
                            return
                else:
                    QMessageBox.warning(self, "File Not Found", f"The file '{fileName}' does not exist.")
                    return
            self.text_animation_timer.start(50)  # Adjust timing for each letter

    def animate_text(self, full_message):
        if self.text_animation_index <= len(full_message):
            self.countdown_label.setText(full_message[:self.text_animation_index])
            self.text_animation_index += 1
        else:
            self.text_animation_timer.stop()

    def start_video_feed(self):
        self.stack.setCurrentWidget(self.video_feed_page)

        self.cap = cv2.VideoCapture(0)
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_feed)
        self.timer.start(10)

    def update_feed(self):

        ret, frame = self.cap.read()

        if ret:
            # Resize frame to aspect fill
            frame_height, frame_width, _ = frame.shape
            target_aspect = self.videoFeedWidth / self.videoFeedHeight
            frame_aspect = frame_width / frame_height

            if frame_aspect > target_aspect:
                new_width = int(frame_height * target_aspect)
                offset = (frame_width - new_width) // 2
                frame = frame[:, offset:offset + new_width]
            else:
                new_height = int(frame_width / target_aspect)
                offset = (frame_height - new_height) // 2
                frame = frame[offset:offset + new_height, :]

            frame = cv2.resize(frame, (self.videoFeedWidth, self.videoFeedHeight))

            # Update video feed
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            bytes_per_line = ch * w
            qimg = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            self.video_feed.setPixmap(QPixmap.fromImage(qimg))

            if self.pause_processing:
                return

            if not self.ready_button_clicked:
                return

            # If not processing, trigger detection
            if not self.processing:
                self.processing = True
                QTimer.singleShot(500, lambda: self.handle_detection(frame))

    def handle_detection(self, frame):
        self.pause_processing = True
        result = self.process_frame(frame)

        if result == "Unknown":
            # Clear labels if no face is detected
            self.id_label.setText("ID: ")
            self.name_label.setText("Name: ")
            self.time_label.setText("Time: ")
            self.date_label.setText("Date: ")

            msg = QMessageBox()
            msg.setWindowTitle("")
            msg.setText("Unknown Face")
            msg.setStandardButtons(QMessageBox.Ok)
            response = msg.exec_()  # Capture the response
            if response == QMessageBox.Ok:
                self.reset_data()
        else:
            # Separate ID and Name from the result
            user_id, user_name = result.split("_", 1)

            self.id_label.setText(f"ID: {user_id}")
            self.name_label.setText(f"Name: {user_name}")

            # Update time and date labels separately
            current_time = time.strftime('%I:%M:%S %p')  # Format time as 12-hour with AM/PM
            current_date = datetime.now().strftime('%Y-%m-%d')

            self.time_label.setText(f"Time: {current_time}")
            self.date_label.setText(f"Date: {current_date}")

            self.id_label.setVisible(True)
            self.time_label.setVisible(True)
            self.date_label.setVisible(True)
            self.name_label.setVisible(True)
            self.attendance_label.setVisible(True)

            if result in self.detected_faces:
                msg = QMessageBox()
                msg.setWindowTitle("Attendance Saved")
                msg.setText("User already marked present.")
                msg.setStandardButtons(QMessageBox.Ok)
                response = msg.exec_()  # Capture the response
                if response == QMessageBox.Ok:
                    self.reset_data()
            else:
                self.detected_faces.add(result)
                self.save_to_spreadsheet(result)

                msg = QMessageBox()
                msg.setWindowTitle("Recognition Success")
                msg.setText("User recognized successfully!")
                msg.setStandardButtons(QMessageBox.Ok)
                response = msg.exec_()  # Capture the response
                if response == QMessageBox.Ok:
                    self.reset_data()

        self.processing = False
        self.pause_processing = False

    def reset_data(self):
        print("OK button was clicked.")
        self.ready_button.setVisible(True)
        self.back_button.setVisible(True)
        self.ready_button_clicked = False
        self.id_label.setVisible(False)
        self.time_label.setVisible(False)
        self.date_label.setVisible(False)
        self.name_label.setVisible(False)
        self.attendance_label.setVisible(False)

    def save_to_spreadsheet(self, result):
        # Define the spreadsheet path
        file_path = "attendance.xlsx"
        # Load or create the workbook
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()

        sheet = workbook.active
        # Check if the headers exist
        if sheet.max_row == 1:
            sheet.append(["ID", "Name", "Date", "Time"])

        # Separate ID and Name from the result
        id, user_name = result.split("_", 1)
        # Get the current date and time
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        # Check if the ID and date already exist in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == id and row[2] == current_date:
                print("Duplicate entry found. Skipping.")
                return

        # Append the data (ID, Name, Date, and Time)
        sheet.append([id, user_name, current_date, current_time])
        # Save the workbook
        workbook.save(file_path)
        # Save data to Realtime DB for the specific class ID
        self.update_realtime_db(id, user_name, current_date, current_time)

    def update_realtime_db(self, student_id, name, date, time):
        ref = db.reference(f'Attendances/{self.class_id}')
        student_data = {
            "name": name,
            "ID": student_id,
            "Date": date,
            "Time": time
        }
        ref.child(student_id).set(student_data)

    def process_frame(self, image):
        # Convert BGR (OpenCV) to RGB (face_recognition library)
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # # Find all face locations in the frame
        face_locations = face_recognition.face_locations(rgb_image)
        encode_cur_frame = face_recognition.face_encodings(rgb_image, face_locations)
        result = "Unknown"

        if face_locations:
            for encodeFace, faceLoc in zip(encode_cur_frame, face_locations):
                matches = face_recognition.compare_faces(self.encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(self.encodeListKnown, encodeFace)

                match_index = np.argmin(faceDis)
                faceDisVal = faceDis[match_index]
                print("Match Index ", match_index,"detectedThres ", faceDisVal)

                # Add a threshold for valid matches
                threshold = 0.4  # Adjust the threshold based on your requirements
                if matches[match_index] and faceDisVal < threshold:
                    self.id = self.studentIds[match_index]
                    top, right, bottom, left = faceLoc
                    cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)

                    # Display text (e.g., 'Face Detected') at the top-left corner of the face
                    result = self.id
                else:
                    print("Face not recognized or below confidence threshold.")
                print(self.id)

        return result

    def open_add_person_page(self):
        class_id = self.class_id_input.text().strip()
        self.add_person_page = AddPersonPage(class_id)
        self.add_person_page.show()

    def closeEvent(self, event):
        if self.cap:
            self.cap.release()
        event.accept()

if __name__ == "__main__":
    app = QApplication([])
    window = FaceRecognitionApp()
    window.show()
    app.exec_()